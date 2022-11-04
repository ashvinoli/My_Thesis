import os
import torch
import numpy as np
import torch.nn as nn
from sklearn.preprocessing import StandardScaler
from sklearn.model_selection import train_test_split
import matplotlib.pyplot as plt
import pandas as pd
import plotly
import plotly.graph_objs as go
from openpyxl import load_workbook
import json
import sys
import re
import shutil
import  glob
import copy
from gen_network_string import gen_string


#Set plot parameters globally
plt.rcParams["font.family"] = "Times New Roman"
plt.rcParams.update({'font.size': 13})
plt.rcParams.update({'font.weight': 'bold'})


class custom_neural_model(nn.Module):
    def __init__(self,input_size,output_size,network_arch) -> None:
        super().__init__()
        self.network_arch = gen_string(network_arch)
        
        #Create sequential network using the layer input from file
        exec(self.network_arch)
       

       

    def forward(self,x):
        return self.linear_regression(x)


class structural_model():
    def __init__(self,file_name,sheet_name,input_range,output_range):
        self.input_file = file_name
        self.sheet_name = sheet_name
        self.input_range = input_range
        self.output_range = output_range
        self.dir_path = os.path.dirname(os.path.realpath(__file__))
        
        
        
    def save_mean_var(self,std_sc,type=True):
        """Saves the mean and var of the input and output data

        Args:
            std_sc (Standard Scalar): standard scalar
            type (bool, optional): if true it is input data i.e X train if false it is output i.e Y train. Defaults to True.
        """
        if type:
            file_name = "./output/input.dat"
        else:
            file_name = "./output/output.dat"

        with open(file_name,"wb") as f:
            #first save means
            np.save(f,std_sc.mean_)
            #then variance
            np.save(f,std_sc.var_)



    def prepare_data(self):
        #read data from excel
        self.df = pd.read_excel(self.input_file,sheet_name=self.sheet_name)

        #Preparing Data Select the required cols
        self.Y_data = self.df.iloc[:,self.output_range]
        Y_data_to_np = self.Y_data.to_numpy()

        self.X_data = self.df.iloc[:,self.input_range]
        X_data_to_np = self.X_data.to_numpy()

        #Break data into test and train data
        X_train,X_test,Y_train,Y_test = train_test_split(X_data_to_np,Y_data_to_np,test_size=0.2,random_state=42)
        self.sc = StandardScaler()

        #Normalize the data
        X_train = self.sc.fit_transform(X_train)

        #While saving true in the second argument indicated X data and False indicates Y data
        self.save_mean_var(self.sc,True)
        X_test = self.sc.transform(X_test)
        
        #Not necessary to transform Y_test as it is of no use to input into the model
        Y_train = self.sc.fit_transform(Y_train)
        self.save_mean_var(self.sc,False)
        #Y_test = self.sc.transform(Y_test)

        #Convert numpy arrays to torch tensors
        X_train = torch.from_numpy(X_train.astype(np.float32))
        Y_train = torch.from_numpy(Y_train.astype(np.float32))
        X_test = torch.from_numpy(X_test.astype(np.float32))
        Y_test = torch.from_numpy(Y_test.astype(np.float32))

        self.X_train = X_train
        self.Y_train = Y_train
        self.X_test = X_test
        self.Y_test = Y_test

    def plot_data(self):
        #Set marker properties
        save_loc = os.path.join(self.dir_path,"output","images_5D")
        markersize = self.X_data[self.X_data.keys()[3]]/6
        markercolor = self.X_data[self.X_data.keys()[1]]

        #Make Plotly figure
        for key in self.Y_data.keys():
            print(f"Plotting for {key}...")
            file_html = os.path.join(save_loc,f"{key}.html")
            file_png = os.path.join(save_loc,f"{key}.png")
            
            gen_data = go.Scatter3d(x=self.X_data[self.X_data.keys()[0]],
                                y=self.X_data[self.X_data.keys()[2]],
                                z=self.Y_data[key],
                                marker=dict(size=markersize,
                                            color=markercolor,
                                            opacity=0.9,
                                            reversescale=True,
                                            colorscale='Viridis'),
                                line=dict (width=0.02),
                                mode='markers')
            fig = go.Figure(data = gen_data)

            #Viewing camera angle
            camera = dict(
                        up=dict(x=0, y=0, z=1),
                        center=dict(x=0
                        , y=0, z=0),
                        eye=dict(x=2, y=0.2, z=0.1)
                        )

            #Make Plot.ly Layout
            mylayout = go.Layout(scene=dict(xaxis=dict( title=self.X_data.keys()[0]),
                                            yaxis=dict( title=self.X_data.keys()[2]),
                                            zaxis=dict(title=key)),scene_camera = camera)
            fig.update_layout(mylayout)

            #Plot and save html
            plotly.offline.plot({"data": [gen_data],
                                "layout": mylayout},
                                auto_open=False,
                                image = 'png', 
                                image_filename='plot_image',
                                output_type='file',
                                image_width=800, 
                                image_height=600,
                                filename=file_html)

            #plot and save png
            fig.write_image(file_png) 

    def get_model_data(self):
        #Extracts linear layers from the model and informs on how many layers and nodes per layer
        str_data = self.model.__str__()
        tokens = str_data.split("\n")
        layers = [line for line in tokens if "Linear" in line]
        nums = [re.findall(r"\d+",item) for item in layers]
        np_arr = np.array(nums)
        mod_arr = np_arr[:,1:]
        all_layers = []
        all_layers.append(mod_arr[0][0])
        all_layers.append(mod_arr[0][1])
        for i in range(1,len(mod_arr)):
            all_layers.append(mod_arr[i,1])

        self.model_data =  all_layers
        str_data = [str(i) for i in self.model_data]
        self.layer_config = '-'.join(str_data)

    def transfer_files(self):
        folder_name = f"Layer {self.layer_config}"
        complete_path = os.path.join(self.dir_path,'output','layers',folder_name)
        os.makedirs(complete_path,exist_ok = True)
        max_dev_path = os.path.join(self.dir_path,"output","max_deviations.txt")
        excel_file_path = os.path.join(self.dir_path,"input","NonLinearData.xlsx")
        trained_model_file = os.path.join(self.dir_path,"output","trained_model.pth")
            
        image_path = os.path.join(self.dir_path,'output','images_regression','*.jpg')
        dat_file = os.path.join(self.dir_path,'output','*.dat')
        for file in glob.glob(image_path):
            shutil.copy2(file,complete_path)

        for file in glob.glob(dat_file):
            shutil.copy2(file,complete_path)

        shutil.copy2(max_dev_path,complete_path)
        shutil.copy2(excel_file_path,complete_path)
        shutil.copy2(trained_model_file,complete_path)

        #Remove excel file if already exists
        if os.path.exists(os.path.join(complete_path,f"{self.layer_config}.xlsx")):
            os.remove(os.path.join(complete_path,f"{self.layer_config}.xlsx"))
        os.rename(os.path.join(complete_path,"NonLinearData.xlsx"),os.path.join(complete_path,f"{self.layer_config}.xlsx"))

    def train_data(self,network_arch):
        learning_rate = 0.001
        momentum = 0.9
        n_iters = 300000
        #Get sizes of input and output data in our case it is 5 and 1
        input_size,output_size = self.X_train.shape[1],self.Y_train.shape[1]
        
        self.model = custom_neural_model(input_size,output_size,network_arch)
        self.get_model_data()
        self.loss_fun = nn.MSELoss() #This class is callable
        self.optimizer = torch.optim.SGD(self.model.parameters(),lr=learning_rate,momentum=momentum)
        loss = ""

        #Save all max deviations, weights and biases and use the one with the lowerst max_deviation
        max_deviations= []
        iterations = []
        model_state_dictionaries = []
        optimizer_state_dictionaries = []
        MSELoss = []

        for epoch in range(n_iters): 
            #forward pass
            Y_pred = self.model(self.X_train)
            #Loss calculation
            loss = self.loss_fun(Y_pred,self.Y_train)
            #back prop
            loss.backward()
            #update weights
            self.optimizer.step()
            #set grads to zero to prevent acculumation
            self.optimizer.zero_grad()

            #print loss every 1000th step but update max deviation every 10th step
            max_deviation = ""
            if ((epoch+1) % 10==0):
                max_deviation = self.test_accuracy(intermediate=True)
                iterations.append(epoch+1)
                max_deviations.append(max_deviation)
                model_state_dictionaries.append(copy.deepcopy(self.model.state_dict()))
                optimizer_state_dictionaries.append(copy.deepcopy(self.optimizer.state_dict()))
                MSELoss.append(loss.item())

            if ((epoch+1) % 1000==0):
                print(f"epoch:{epoch+1} Loss:{loss.item():.4f} Max Deviation:{max_deviation:.4f}")

        min_index = max_deviations.index(min(max_deviations))
        #print(min_index)
        #load the model with minimum error and loss
        self.model.load_state_dict(model_state_dictionaries[min_index])
        self.optimizer.load_state_dict(optimizer_state_dictionaries[min_index])
        self.MSELoss = MSELoss[min_index]

        plt.plot(iterations,max_deviations)
        plt.xlabel("No. of Iterations")
        plt.ylabel("Maximum Deviation(%)")
        plt.grid()
        plt.savefig(f'./output/images_regression/output_iterations_vs_max_deviation.jpg',bbox_inches='tight')
        plt.clf()

    def test_accuracy(self,intermediate=False):
        """This will test the accuracy of the model by finding the difference in datavalues. It will also write the Y_pred and Y_test to a excel file along with its difference
        """
        if not intermediate:
            print("Testing Accuracy...")
        #We don't want torch to keep updating its computational graph so torch.nograd is necessary
        with torch.no_grad():
            #The predicted outcome will be the normalized one as the model is trained on normalized data, so inverse transform has to be done in the predicted y data. The mean and sd required for inverse transform is obtained from the training data Y values i.e Y_train

            Y_test_predicted = self.model(self.X_test)
            inversed_Y = self.sc.inverse_transform(Y_test_predicted)

            input_and_output = list(zip(self.Y_test.numpy(),inversed_Y))
            #print(input_and_output)

            #inversed are the data obtained from neural net after reverse normalizing
            inversed_Y_Dataframe = pd.DataFrame(inversed_Y,columns=self.Y_data.keys())
            actual_Y_dataframe = pd.DataFrame(self.Y_test.numpy(),columns=self.Y_data.keys())
            Differentiated_data_frame = pd.DataFrame()
            data_frame_with_both_data = pd.DataFrame()
            top_line_str = ""
            max_diffs = []
            for key in self.Y_data.keys():
                col_name = f"Difference_{key}"
                col_name_percent = f"Difference_{key}_Percent"
                Differentiated_data_frame[col_name] = abs(inversed_Y_Dataframe[key] - actual_Y_dataframe[key])
                data_frame_with_both_data["Initial"] = actual_Y_dataframe[key].values
                data_frame_with_both_data["Final"] = inversed_Y_Dataframe[key].values
                Differentiated_data_frame[col_name_percent] = abs(inversed_Y_Dataframe[key] - actual_Y_dataframe[key])/actual_Y_dataframe[key] * 100

                top_line_str += f"\t{col_name_percent}"
                max_diffs.append(Differentiated_data_frame[col_name_percent].max())

                if not intermediate:           
                    #plot
                    ax = data_frame_with_both_data.plot(kind= 'scatter',x="Initial",y="Final")
                    
                    #get regression params
                    d = np.polyfit(data_frame_with_both_data['Initial'],data_frame_with_both_data['Final'],1)
                    f= np.poly1d(d)
                    data_frame_with_both_data["Best Fit"] = f(data_frame_with_both_data["Initial"]
                    )
                    data_frame_with_both_data.plot(x="Initial",y="Best Fit",color = "Red",ax=ax)
                    ax.set_xlabel(f"{key} Predicted by Ansys")
                    ax.set_ylabel(f"{key} Predicted by Neural Net")
                    plt.legend([f"y={f.coefficients[0]:.3f}x+{f.coefficients[1]:.4f}"])
                    print(f"m={f.coefficients[0]} and c = {f.coefficients[1]}")
                    #ax.text(200,310,f"y={f.coefficients[0]:.3f}x+{f.coefficients[1]:.3f}")
                    #plt.show()
                    plt.grid()
                    plt.savefig(f'./output/images_regression/output_{key}.jpg',bbox_inches='tight')
                    plt.clf()

            #if testing is done intermediately return maximum of max_diffs
            if intermediate:
                return max(max_diffs)

            
            file_empty = False
            self.max_deviations = open("./output/max_deviations.txt","r")
            lines =list(self.max_deviations.readlines())
            final_lines = [i for i in lines if i is not '\n'] 
            if len(final_lines)==0:
                file_empty = True
            self.max_deviations.close()

            if file_empty:
                self.max_deviations = open("./output/max_deviations.txt","w")
                self.max_deviations.write(f"Network Model \t MSELoss \t {top_line_str}\n")
            else:
                self.max_deviations = open("./output/max_deviations.txt","a")
            data_str = [str(i) for i in max_diffs]
            insert_data = "\t".join(data_str)
            insert_str = f"{self.layer_config} \t {self.MSELoss:.4f} \t {insert_data}\n"
            self.max_deviations.write(insert_str)
            self.max_deviations.close()

            #Write real and predicted outcome data to excel. Following way appends the data to the existing excel file
            
            print("Saving Excel File...")
            book = load_workbook(self.input_file)
            all_sheets_name = book.sheetnames
            added_sheet_names = ['Output_Inversed_Y','Output_Actual_Y','Output_Y_Difference']
            for sheet_name in added_sheet_names:
                if sheet_name in all_sheets_name:
                    book.remove(book[sheet_name])


            writer = pd.ExcelWriter(self.input_file, engine = 'openpyxl')
            writer.book = book

            inversed_Y_Dataframe.to_excel(writer, sheet_name = 'Output_Inversed_Y')
            actual_Y_dataframe.to_excel(writer, sheet_name = 'Output_Actual_Y')
            Differentiated_data_frame.to_excel(writer, sheet_name = 'Output_Y_Difference')
            writer.save()
            writer.close()

    def save_trained_model(self):
         torch.save(self.model, "./output/trained_model.pth")

         #After all training has been done transfer all necessary files to respective folder
         print("Copying files to another folder...")
         self.transfer_files()
         print("All Done!")

    def load_trained_model(self,list_of_inputs=None):
        head = "./output/layers"
        all_dirs = [f"{head}/{folder.name}" for folder in os.scandir("./output/layers")]
        for item in all_dirs:
            complete_file = os.path.join(self.dir_path,item,"trained_model.pth")
            if os.path.exists(complete_file):
                self.loaded_model = torch.load(complete_file)
                self.loaded_model.eval()

                #read input_data from json file if list_of_inputs is not given else set it to list_of_inputs
                data = ""
                if list_of_inputs:
                    data = list_of_inputs
                else:
                    with open("./input/input_data.json") as f:
                        data = json.load(f)["input_data"]

                #read mean and sd for input and output. the read data is a numpy array consisting of the means and variances of the data used to train the neural net.
                
                X_file = open(os.path.join(self.dir_path,item,"input.dat"),"rb")
                Y_file = open(os.path.join(self.dir_path,item,"output.dat"),"rb")
                X_mean = np.load(X_file)
                X_var = np.load(X_file)
                Y_mean = np.load(Y_file)
                Y_var = np.load(Y_file)

                #output = []
                print(f"For {item}:")
                for item in data:
                    #first normalize the read data
                    normalized_input = (item-X_mean)/X_var**(1/2)
                    tensor_input = torch.from_numpy(normalized_input.astype(np.float32))
                    #The following line is essential because we are forecasting data and don't want torch to track gradients
                    with torch.no_grad():
                        normalized_tensor_output = self.loaded_model(tensor_input)
                        normalized_np_output = normalized_tensor_output.numpy()
                        real_output = normalized_np_output*Y_var**(1/2)+Y_mean
                        #output.append(real_output)
                        print(real_output)

                X_file.close()
                Y_file.close()
                print("\n")
                #return output
            else:
                print(f"Model not available for {item}")


def main_screen():
    screen_string = """
    1. Train, Test and save Model
    2. Load Trained Model and Predict
    3. Visualize Data
    4. Exit
    """
    print(screen_string)
    
def main():
    #Slice Object acts as 0:4 and 4: for input to numpy slicing
    sheet = ""
    with open("./input/input_sheet.json") as f:
        data = json.load(f)
        sheet = data["sheet_name"]


    my_model = structural_model('./input/NonLinearData.xlsx',sheet,slice(0, 5, None),slice(5, None, None))

    def choice_one():
        with open("./input/input_batch_network_layers.json") as f:
                data = json.load(f)
                layers = data["Layers"]
                for item in layers:
                    print(f"Working for Layer {item}\n\n")
                    my_model.prepare_data()
                    my_model.train_data(item)
                    my_model.test_accuracy()
                    my_model.save_trained_model()

    if len(sys.argv) == 1:
        main_screen()
        choice = int(input("Enter your choice: "))
        if choice == 1:
           choice_one()
        elif choice == 2:
            my_model.load_trained_model()
        elif choice == 3:
            my_model.plot_data()
        else:
            print("Invalid choice")
    else:
        if sys.argv[1]=='p':
            my_model.load_trained_model()
        elif sys.argv[1]=='t':
            choice_one()
        else:
            print("Invalid Command. Quitting!")    


if __name__ == '__main__':
    main()